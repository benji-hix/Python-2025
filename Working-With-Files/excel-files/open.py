"""Read/Write Excel File"""

from openpyxl import Workbook, load_workbook
from pathlib import Path

wkbk_name = ''

def clean_str(str: str) -> str:
    """
    Removes enclosing quote marks from str.

    Args:
        str (str): String to clean.

    Returns:
        str: Text-only string.
    """
    return str.strip("'").strip('"')

def get_dir_path() -> Path:
    """
    Receive user input to select target directory.

    Raises:
        FileNotFoundError: Invalid directory path.

    Returns:
        Path: Target directory's path object.
    """
    while True:
        try:
            dir_str = clean_str(input('\n| Paste directory path:\n| '))
            target_dir = Path(dir_str)
            if not target_dir.is_dir():
                raise FileNotFoundError(f'| {dir_str} not found.')
            else:
                print('\n| Directory loaded successfully.')
                break
        except FileNotFoundError as e:
            print(f'| Error: {e}')

    return target_dir

def print_dir_files(dir: Path) -> None:
    """
    Print all .xlsx files in a directory.

    Args:
        dir (Path): Directory path object.

    Returns:
        None
    """
    dir_files = [
        f for f in dir.iterdir() 
        if f.is_file() and f.suffix == '.xlsx' and not f.name.startswith('~$')
        ] 
    if not dir_files:
        print('| No files found.')
    else:
        print('| Available files:')
        for f in dir_files:
            print(f'|····{f.name}')
    return None

def get_wkbk_path(dir: Path) -> Path:
    """
    Receives user input to select a workbook from
    available .xlsx files in a directory.

    Args:
        dir (Path): Directory path object.

    Raises:
        FileNotFoundError: Invalid file name.

    Returns:
        Path: Target file's path object.
    """
    global wkbk_name
    print_dir_files(dir)

    while True:
        try:
            wkbk_name = clean_str(input('\n| Select file:\n| '))
            if not wkbk_name.endswith('.xlsx'):
                wkbk_name += '.xlsx'
            wkbk_path = dir / wkbk_name
            if not wkbk_path.exists():
                raise FileNotFoundError(f'| {wkbk_name} not found.')
            else:
                break
        except FileNotFoundError as e:
            print(f'| Error: {e}')

    return wkbk_path

def open_wkbk(wkbk_path: Path) -> Workbook:
    """
    Open target file's workbook.

    Args:
        wkbk_path (Path): Target file path object.

    Returns:
        Workbook: Target file Workbook object.
    """
    print('\n| File loaded successfully.')
    return load_workbook(str(wkbk_path))

def print_wksheets(wkbk: Workbook) -> None:
    """
    Print all worksheets available in target Workbook.

    Args:
        wkbk (Workbook): Workbook object.

    Returns:
        None
    """
    print('| Available worksheets:')
    for s in wkbk.sheetnames:
        print(f'|····{s}')
    return None

def open_wksheet(wkbk_path: Path) -> Workbook:
    """
    Select and open one worksheet from a workbook.

    Args:
        wkbk_path (Path): Workbook object.

    Raises:
        FileNotFoundError: Invalid worksheet name.

    Returns:
        Workbook: Active worksheet.
    """
    current_wkbk = open_wkbk(wkbk_path)
    print_wksheets(current_wkbk)

    while True:
        try:
            sheet_name = clean_str(input('\n| Select worksheet:\n| '))
            if sheet_name not in current_wkbk.sheetnames:
                raise FileNotFoundError(f'{sheet_name} not found.')
            else:
                sheet = current_wkbk[f'{sheet_name}']
                print('\n| Spreadsheet loaded successfully.')
                break
        except FileNotFoundError as e:
            print(f'| Error: {e}')
        
    return sheet

def open_spreadsheet() -> Workbook:
    """
    Receive user input to access 
    correct spreadsheet and await read/update.

    Args:
        None

    Returns:
        Workbook: Active spreadsheet.
    """

    dir_path = get_dir_path()
    wkbk_path = get_wkbk_path(dir_path)
    current_wksht = open_wksheet(wkbk_path) 

    return current_wksht