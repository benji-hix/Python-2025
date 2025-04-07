from openpyxl import Workbook, load_workbook

def parse_row_values(row: list) -> dict:
    row_dict = {
        'Serial': row[0],
        'Person ID' : row[1],
        'Name' : row[2],
        'Date' : row[3],
        'Case' : row[4],
        'Lost Mode' : row[5],
        'Status' : row[6]
    }
    return row_dict

def print_row(current_wksht):
    # current_wkbk = load_workbook(r'D:\benji\code\excel-testing.xlsx')
    # current_wksht = current_wkbk['Open-Cases']

    selected_row = int(input('| Select a row:\n| '))
    # row_list = list(next(current_wksht.iter_rows(selected_row, selected_row, values_only=True)))

    row_dict = parse_row_values(next(current_wksht.iter_rows(selected_row, selected_row, values_only=True)))

    for key, value in row_dict.items():
        print(f'| {key} : {value}')

